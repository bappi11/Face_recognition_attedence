import cv2
import pickle
import openpyxl as xl
import datetime as dt
import numpy as np

cap = cv2.VideoCapture(0)
def update_data(present_std):
    wb = xl.load_workbook('attendence.xlsx')
    sheet = wb['Sheet1']
    t_dt = sheet.max_column
    t_dt += 1

    cell3 = sheet.cell(1, t_dt)
    today = dt.date.today()
    print(today)
    cell3.value = dt.date.today()
    for i in range(2, sheet.max_row + 1):
        cell1 = sheet.cell(i, 1)
        cell3 = sheet.cell(i, t_dt)
        if cell1.value in present_std:
            cell3.value = 'P'
        else:
            cell3.value = 'A'

    wb.save("attendence.xlsx")



face_cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read("trainner.yml")
labels = {}
present_id = []
with open("labels.pikle", 'rb') as f:

    og_labels = pickle.load(f)
    labels = {v: k for k, v in og_labels.items()}

while True:
    ret, frame = cap.read()
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.3, 5)

    for (x, y, w, h) in faces:
        roi_gray = gray[y:y + h, x:x + w]
        cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 0, 0), 4)
        id_, conf = recognizer.predict(roi_gray)
        font = cv2.FONT_HERSHEY_SIMPLEX
        if conf >= 45 and conf <= 85:
            print(id_)
            print(labels[id_])
            name = labels[id_]
            f_name = f"{name} \n {int(conf)}"
            color = (255, 255, 0)
            cv2.putText(frame, f_name, (x, y), font, 1, color, 2)
            if name not in present_id:
                present_id.append(name)

        else:
            f_name = f" Unknown {int(conf)}"
            color = (255, 255, 0)
            cv2.putText(frame, f_name, (x, y), font, 1, color, 2)

    cv2.imshow('Video', frame)
    k = cv2.waitKey(1)
    if k == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()

print(present_id)
update_data(present_id)
